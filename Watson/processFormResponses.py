# -*- coding: utf-8 -*-
#import json
from watson_developer_cloud import ConversationV1
from openpyxl import load_workbook
import re

#%% reading existing examples
reAllLettersAndSpace = u"[A-Za-záàãâçéêíóôúü\s]"
filename = u"../Dados para Chatbot do Livro dos Espíritos.xlsx"

def get_q_number(q_text):
    return re.findall("^(.+?). ", q_text)[0]

def process_example(q_marked):
    text = q_marked
#    print(text)
    terms = re.findall("\["+reAllLettersAndSpace+"+\s?@?\w*\]",text)
    for term in terms:
        term_text = term[1:-1]
        term_text = re.split("\s+@", term_text)[0]
        text = text.replace(term,term_text,1)
    return text

wb = load_workbook(filename, keep_vba = True)
s_qa = wb.get_sheet_by_name("Perguntas e Respostas")
example_dict = dict()
qa_row_dict = dict()
for row_num in range(2,s_qa.max_row+1):
    row = s_qa[row_num]
    q = row[0].value
    if q is None:
        break
    q_num = get_q_number(q)
    qa_row_dict[q_num] = row_num
    # pending: remove tagging
    example_dict[q_num] = [process_example(cell.value) for cell in row[3:] if cell.value is not None]

#%% reading and processing new data from form resposes
s_form = wb.get_sheet_by_name("Respostas Form")
new_example_dict = dict()

# initializing watson parameters
conversation = ConversationV1(
    username='8fcdf6b6-be15-400b-bb6d-1d5e2c44c2d3',
    password='lieR8OpYbMKV',
    version='2017-02-03')

workspace_id = '87d42987-eec9-4427-a63a-fc1e071b2f2b'
examples_added_count = 0
for row in s_form.rows:
    processed = (row[4].value is not None)
    if processed:
        continue
    q_num = row[1].value
    if type(q_num) in (float,int):
        q_num = "%d"%q_num
    else:
        q_num = q_num.lower().replace(" ","")
    text = row[2].value
    print("\nNew example for question %s: %s"%(q_num,text))
    if text in example_dict[q_num]:
        print("Example is already in the spreadsheet.")
        row[4].value = "Processado"
        continue
    example_count = len(example_dict[q_num])
    if example_count == 0:
        print("This question has not been evaluated yet.")
    else:
        response = conversation.message(workspace_id=workspace_id, message_input={
        'text': text})
        response_text = response["output"]["text"][0]
        print("Response from Watson: %s"%response_text)
    #    print(json.dumps(response, indent=2))
    #    print(response["output"]["text"][0])
        r_num = get_q_number(response_text)
        if r_num == q_num: # watson has correctly identified the question
            row[4].value = "Processado"
            continue
        # inserting new example to spreadsheet
    print("-> Adding example to spreadsheet.")
    qa_row_num = qa_row_dict[q_num]
    qa_row = s_qa[qa_row_num]
    example_idx = len(example_dict[q_num]) + 3
    qa_row[example_idx].value = text
    example_dict[q_num].append(text)
    row[4].value = "Processado"
    examples_added_count += 1

print("\nAdded %d examples to the spreadsheet."%examples_added_count)
wb.save(filename)

        
        
        
        
        
        
        
        
        
        
        
        
        
        
        